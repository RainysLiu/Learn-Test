import openpyxl
import pypinyin

# 读取excel数据
def read_excel(excel_path, sheet_name):
    work_book = openpyxl.load_workbook(excel_path)
    if work_book:
        sheet = work_book[sheet_name]
    else:
        sheet = work_book.active
    data = []
    for row in sheet.rows:
        row_list = [cell.value for cell in row]
        data.append(row_list)
    # print(data)
    return data

# 重新写入excel数据
def rewrite_excel(excel_path, sheet_name, data):
    work_book = openpyxl.load_workbook(excel_path) # 表存在，导入
    sheet = work_book.get_sheet_by_name(sheet_name)
    # enumerate() # 将一个可遍历的数据对象(如列表、元组或字符串)组合为一个索引序列
    print(list(enumerate(data))) # [(0,row1_data),(1,row2_data),....]

    for row, item in enumerate(data):
        for co1, value in enumerate(item):
            sheet.cell(row+1, co1+1, value)

    work_book.save("my_excel_test.xlsx")


# 往原有sheet 表里加内容
def append_excel(excel_path, sheet_name):
    work_book = openpyxl.load_workbook(excel_path)
    sheet = work_book.get_sheet_by_name(sheet_name)
    nrows = sheet.max_row
    print("原来总行数为",nrows)
    nco1s = sheet.max_column
    append_data = [[1111, "张三", "12345678900", " ", " ", " "],
                   [2222, "李四", "13992515990"]]

    # 方法一: 直接append
    # for data in append_data:
    #     sheet.append(data)


    # 方法二: 同上一函数
    # for row, item in enumerate(append_data):
    #     for co1, value in enumerate(item):
    #         sheet.cell(row + nrows + 1, co1 + 1, value)

     # 方法三 :
    for row in range(nrows+1, nrows+len(append_data)+1):
        for col in range(1, len(append_data[row-nrows-1])+1):
            sheet.cell(row, col).value = append_data[row-nrows-1][col-1]


    nrows = sheet.max_row
    print("现在总行数为",nrows)

    work_book.save("my_append_test.xlsx")

def add_cols_pinyin(data):
    for i, row_data in enumerate(data[1:]):
        # style=pypinyin.NORMAL 不显示声调
        name_pinyin = pypinyin.pinyin(row_data[1], style=pypinyin.NORMAL)
        # print(name_pinyin)
        name_pinyin_suoxie =''.join([d[0][0] for d in name_pinyin[:2]])
        row_data.insert(2,name_pinyin_suoxie)
        # print(row_data)
    data = sorted(data[1:], key=lambda x : x[2])
    print(data)

def sort_ID_excel(data):
    res = sorted(data[1:], key = lambda x: x[0])
    print(res)



if __name__ == '__main__':
    data = read_excel("5.62.xlsx","Worksheet")
    rewrite_excel("my_test.xlsx","Worksheet",data)
    # append_excel("my_test.xlsx","Worksheet")
    # sort_ID_excel(data)
    add_cols_pinyin(data)

