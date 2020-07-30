# coding = utf-8
import openpyxl
from openpyxl.styles import Alignment  # excel格式设置
import pymysql


# 读取MySQL db_movie 数据库数据
def con_mysql():
    con = pymysql.Connect(host="localhost", port=3306,
                          user="root", passwd="123456",
                          db="db_movie", charset="utf8",
                          )
    # con = pymysql.connections.Connection(cursorclass=pymysql.cursors.DictCursor,
    #                                      host="localhost", port=3306,
    #                                      user="root", passwd="123456",
    #                                      db="db_movie", charset="utf8",
    #                                      )
    cur = con.cursor()
    sql1 = "use db_movie;"
    cur.execute(sql1)
    sql2 = "show tables;"
    cur.execute(sql2)
    sql3 = "select * from movie;"
    cur.execute(sql3)
    data = cur.fetchall()
    print(data)
    sql4 = "select COLUMN_NAME from INFORMATION_SCHEMA.Columns where table_name='movie' " \
           "and table_schema='db_movie';"
    cur.execute(sql4)
    excel_title = cur.fetchall()
    print(excel_title)
    return data, excel_title


def excel_mysql(excel_path, excel_title, data):
    work_book = openpyxl.Workbook()  # 表不存在，创建
    sheet1 = work_book.active
    sheet1.title = "movie"
    max_row = sheet1.max_row
    max_col = sheet1.max_column
    sheet1.alignment = Alignment(horizontal='center', vertical='center')
    # sheet1 = work_book.create_sheet("movie", 0)
    for col in range(1, len(excel_title)+1):
        sheet1.cell(1, col, excel_title[col-1][0])

    for row in range(2, len(data) + 2):
        for col in range(1, len(excel_title)+1):
            sheet1.cell(row, col).value = data[row-2][col-1]

    # 设置单元格内容居中
    for row in sheet1.iter_rows(max_row, max_col):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


    work_book.save(excel_path)


if __name__ == '__main__':
    data, excel_title = con_mysql()
    excel_mysql("mysql_excel.xlsx", excel_title, data)