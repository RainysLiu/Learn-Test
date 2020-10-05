import openpyxl


def get_all_text_input():
    """
    获取测试用例信息
    :return:
    """
    # 将测试用例信息存在一个excel表格中，读取excel表格获取内容
    work_book = openpyxl.load_workbook("cart_test.xlsx")
    sheet = work_book['Sheet1']
    all_input_data = []
    for row in range(2, sheet.max_row + 1):
        all_input_data.append([sheet.cell(row, 2).value, sheet.cell(row, 3).value])
    return all_input_data


if __name__ == '__main__':
    get_all_text_input()
